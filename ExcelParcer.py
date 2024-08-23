import openpyxl
def dataAnalysisParser ():
    wb = openpyxl.load_workbook('ArchaeaGenomeProject.xlsx')
    sheet = wb.active
    last_row=sheet.max_row
    # Parse through the results and identify which specific clades were being compared 
    for specific_Row in range (1,last_row+1):
        # access the results which should be in the first coulmn
        list_of_values = sheet.cell(specific_Row,1).value
        data = list_of_values.split()
        # matchtospecific clade is a method that takes in a number and matches it to the specific group
        # only the first 2 numbers matter
        firstSpecificGroup = matchToSpecificClade(data[0])
        secondSpecificGroup = matchToSpecificClade(data[1])
        # modify the next column with the rewritten name
        sheet.cell(specific_Row,2).value = f'{firstSpecificGroup} {secondSpecificGroup}'
        # save
        wb.save('ArchaeaGenomeProject.xlsx')

    # Take the results in this new column and now input in the major clade
    for initialCompareRow in range (1,last_row+1):
        # Essentially doing the same thing as the previous row
        comparision = sheet.cell(initialCompareRow,2).value
        data = comparision.split()
        # This time we are matching the name of the phyla with major clade
        firstMajorGroup = matchToLargeClade(data[0])
        secondMajorGroup = matchToLargeClade(data[1])
        sheet.cell(initialCompareRow,3).value = f'{firstMajorGroup} {secondMajorGroup}'
        wb.save('ArchaeaGenomeProject.xlsx')

    
    # Get the overall count information for all signficant results for each intercomparisons
    Methanomada_Acherontia = 0
    Methanomada_Stygia = 0
    Methanomada_Proteoarchaeaota = 0
    Methanomada_Diaforarchaea = 0
    Methanomada_Methanotecta = 0
    Acherontia_Stygia = 0
    Acherontia_Proteoarchaeaota = 0
    Acherontia_Diaforarchaea = 0
    Acherontia_Methanotecta = 0
    Stygia_Proteoarchaeaota = 0
    Stygia_Diaforarchaea = 0 
    Stygia_Methanotecta = 0
    Proteoarchaeaota_Diaforarchaea = 0
    Proteoarchaeaota_Methanotecta = 0
    Diaforarchaea_Methanotecta = 0
    for final_analysis_row in range (1,last_row):
        #once again accessing the information
        list_of_values = sheet.cell(final_analysis_row,1).value
        data = list_of_values.split()
        # sixth datapoint is either true or false
        if data[6]=='True':
            # get the major clade information and then add a counter to the approiate bucket
            majorGroup  = sheet.cell(final_analysis_row,3).value
            if majorGroup == 'Methanomada Acherontia' or majorGroup == 'Acherontia Methanomada':
                Methanomada_Acherontia = Methanomada_Acherontia + 1
            elif majorGroup == 'Methanomada Stygia' or majorGroup ==  'Stygia Methanomada':
                Methanomada_Stygia = Methanomada_Stygia + 1
            elif majorGroup == 'Methanomada Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Methanomada':
                Methanomada_Proteoarchaeaota = Methanomada_Proteoarchaeaota + 1
            elif majorGroup == 'Methanomada Diaforarchaea' or majorGroup == 'Diaforarchaea Methanomada':
                Methanomada_Diaforarchaea = Methanomada_Diaforarchaea + 1
            elif majorGroup == 'Methanomada Methanotecta' or majorGroup == 'Methanotecta Methanomada':
                Methanomada_Methanotecta = Methanomada_Methanotecta + 1
                
            elif majorGroup == 'Acherontia Stygia' or majorGroup == 'Stygia Acherontia':
                Acherontia_Stygia = Acherontia_Stygia + 1
            elif majorGroup == 'Acherontia Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Acherontia':
                Acherontia_Proteoarchaeaota = Acherontia_Proteoarchaeaota + 1
            elif majorGroup == 'Acherontia Diaforarchaea' or majorGroup == 'Diaforarchaea Acherontia':
                Acherontia_Diaforarchaea = Acherontia_Diaforarchaea + 1
            elif majorGroup == 'Acherontia Methanotecta' or majorGroup == 'Methanotecta Acherontia':
                Acherontia_Methanotecta = Acherontia_Methanotecta + 1

            elif majorGroup == 'Stygia Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Stygia':
                Stygia_Proteoarchaeaota = Stygia_Proteoarchaeaota + 1
            elif majorGroup == 'Stygia Diaforarchaea' or majorGroup == 'Diaforarchaea Stygia':
                Stygia_Diaforarchaea = Stygia_Diaforarchaea + 1
            elif majorGroup == 'Stygia Methanotecta' or majorGroup == 'Methanotecta Stygia':
                Stygia_Methanotecta = Stygia_Methanotecta + 1

            elif majorGroup == 'Proteoarchaeaota Diaforarchaea' or majorGroup == 'Diaforarchaea Proteoarchaeaota':
                Proteoarchaeaota_Diaforarchaea = Proteoarchaeaota_Diaforarchaea + 1
            elif majorGroup == 'Proteoarchaeaota Methanotecta' or majorGroup == 'Methanotecta Proteoarchaeaota':
                Proteoarchaeaota_Methanotecta = Proteoarchaeaota_Methanotecta + 1

            elif majorGroup == 'Diaforarchaea Methanotecta' or majorGroup == 'Methanotecta Diaforarchaea':
                Diaforarchaea_Methanotecta = Diaforarchaea_Methanotecta + 1
    # Once the counting is done in a new column and row, save the information in each bucket along with the comparing specific group
    sheet.cell(1,4).value = 'Methanomada-Acherontia'
    sheet.cell(1,5).value = Methanomada_Acherontia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(2,4).value = 'Methanomada-Stygia'
    sheet.cell(2,5).value = Methanomada_Stygia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(3,4).value = 'Methanomada-Proteoarchaeaota'
    sheet.cell(3,5).value = Methanomada_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(4,4).value = 'Methanomada-Diaforarchaea'
    sheet.cell(4,5).value = Methanomada_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(5,4).value = 'Methanomada-Methanotecta'
    sheet.cell(5,5).value = Methanomada_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(6,4).value = 'Acherontia-Stygia'
    sheet.cell(6,5).value = Acherontia_Stygia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(7,4).value = 'Acherontia-Proteoarchaeaota'
    sheet.cell(7,5).value = Acherontia_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(8,4).value = 'Acherontia-Diaforarchaea'
    sheet.cell(8,5).value = Acherontia_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(9,4).value = 'Acherontia-Methanotecta'
    sheet.cell(9,5).value = Acherontia_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(10,4).value = 'Stygia-Proteoarchaeaota'
    sheet.cell(10,5).value = Stygia_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(11,4).value = 'Stygia-Diaforarchaea'
    sheet.cell(11,5).value = Stygia_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(12,4).value = 'Stygia-Methanotecta'
    sheet.cell(12,5).value = Stygia_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(13,4).value = 'Proteoarchaeaota-Diaforarchaea'
    sheet.cell(13,5).value = Proteoarchaeaota_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(14,4).value = 'Proteoarchaeaota-Methanotecta'
    sheet.cell(14,5).value = Proteoarchaeaota_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(15,4).value = 'Diaforarchaea-Methanotecta'
    sheet.cell(15,5).value = Diaforarchaea_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')


   # Same thing as described previously but this time we ignore the True statement as we are counting the toal number of comparisons
    Methanomada_Acherontia = 0
    Methanomada_Stygia = 0
    Methanomada_Proteoarchaeaota = 0
    Methanomada_Diaforarchaea = 0
    Methanomada_Methanotecta = 0
    Acherontia_Stygia = 0
    Acherontia_Proteoarchaeaota = 0
    Acherontia_Diaforarchaea = 0
    Acherontia_Methanotecta = 0
    Stygia_Proteoarchaeaota = 0
    Stygia_Diaforarchaea = 0 
    Stygia_Methanotecta = 0
    Proteoarchaeaota_Diaforarchaea = 0
    Proteoarchaeaota_Methanotecta = 0
    Diaforarchaea_Methanotecta = 0
    for final_analysis_row in range (1,last_row):
        list_of_values = sheet.cell(final_analysis_row,1).value
        data = list_of_values.split()
        majorGroup  = sheet.cell(final_analysis_row,3).value
        if majorGroup == 'Methanomada Acherontia' or majorGroup == 'Acherontia Methanomada':
            Methanomada_Acherontia = Methanomada_Acherontia + 1
        elif majorGroup == 'Methanomada Stygia' or majorGroup ==  'Stygia Methanomada':
            Methanomada_Stygia = Methanomada_Stygia + 1
        elif majorGroup == 'Methanomada Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Methanomada':
            Methanomada_Proteoarchaeaota = Methanomada_Proteoarchaeaota + 1
        elif majorGroup == 'Methanomada Diaforarchaea' or majorGroup == 'Diaforarchaea Methanomada':
            Methanomada_Diaforarchaea = Methanomada_Diaforarchaea + 1
        elif majorGroup == 'Methanomada Methanotecta' or majorGroup == 'Methanotecta Methanomada':
            Methanomada_Methanotecta = Methanomada_Methanotecta + 1
            
        elif majorGroup == 'Acherontia Stygia' or majorGroup == 'Stygia Acherontia':
            Acherontia_Stygia = Acherontia_Stygia + 1
        elif majorGroup == 'Acherontia Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Acherontia':
            Acherontia_Proteoarchaeaota = Acherontia_Proteoarchaeaota + 1
        elif majorGroup == 'Acherontia Diaforarchaea' or majorGroup == 'Diaforarchaea Acherontia':
            Acherontia_Diaforarchaea = Acherontia_Diaforarchaea + 1
        elif majorGroup == 'Acherontia Methanotecta' or majorGroup == 'Methanotecta Acherontia':
            Acherontia_Methanotecta = Acherontia_Methanotecta + 1

        elif majorGroup == 'Stygia Proteoarchaeaota' or majorGroup == 'Proteoarchaeaota Stygia':
            Stygia_Proteoarchaeaota = Stygia_Proteoarchaeaota + 1
        elif majorGroup == 'Stygia Diaforarchaea' or majorGroup == 'Diaforarchaea Stygia':
            Stygia_Diaforarchaea = Stygia_Diaforarchaea + 1
        elif majorGroup == 'Stygia Methanotecta' or majorGroup == 'Methanotecta Stygia':
            Stygia_Methanotecta = Stygia_Methanotecta + 1

        elif majorGroup == 'Proteoarchaeaota Diaforarchaea' or majorGroup == 'Diaforarchaea Proteoarchaeaota':
            Proteoarchaeaota_Diaforarchaea = Proteoarchaeaota_Diaforarchaea + 1
        elif majorGroup == 'Proteoarchaeaota Methanotecta' or majorGroup == 'Methanotecta Proteoarchaeaota':
            Proteoarchaeaota_Methanotecta = Proteoarchaeaota_Methanotecta + 1

        elif majorGroup == 'Diaforarchaea Methanotecta' or majorGroup == 'Methanotecta Diaforarchaea':
            Diaforarchaea_Methanotecta = Diaforarchaea_Methanotecta + 1


    sheet.cell(1,6).value = Methanomada_Acherontia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(2,6).value = Methanomada_Stygia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(3,6).value = Methanomada_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(4,6).value = Methanomada_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(5,6).value = Methanomada_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(6,6).value = Acherontia_Stygia
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(7,6).value = Acherontia_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(8,6).value = Acherontia_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(9,6).value = Acherontia_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(10,6).value = Stygia_Proteoarchaeaota
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(11,6).value = Stygia_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(12,6).value = Stygia_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(13,6).value = Proteoarchaeaota_Diaforarchaea
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(14,6).value = Proteoarchaeaota_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

    sheet.cell(15,6).value = Diaforarchaea_Methanotecta
    wb.save('ArchaeaGenomeProject.xlsx')

# Just mapping out the index with the specific clade
def matchToSpecificClade (value):
    index = int(value)
    if index == 0:
        return 'Methanococci'
    elif index == 1:
        return 'Methanopyri'
    elif index == 2:
        return 'Thermococci'
    elif index == 3:
        return 'Theionarchaea'
    elif index == 4:
        return 'Hadarchaeia'
    elif index == 5:
        return 'Odinarchaeota'
    elif index == 6:
        return 'Wukongarchaeota'
    elif index == 7:
        return 'Freyarchaeota'
    elif index == 8:
        return 'Baldrarchaeota'
    elif index == 9:
        return 'Heimdallarchaeota'
    elif index == 10:
        return 'Helarchaeota'
    elif index == 11:
        return 'Kariarchaeota'
    elif index == 12:
        return 'Lokiarchaeota'
    elif index == 13:
        return 'Sifarchaeota'
    elif index == 14:
        return 'Thorarchaeota'
    elif index == 15:
        return 'Hodarchaeota'
    elif index == 16:
        return 'Hermodarchaeota'
    elif index == 17:
        return 'Brockarchaeota'
    elif index == 18:
        return 'Culexarchaeota'
    elif index == 19:
        return 'Geoarchaeota'
    elif index == 20:
        return 'Geothermarchaeota'
    elif index == 21:
        return 'Korarchaeota'
    elif index == 22:
        return 'Marsarchaeota'
    elif index == 23:
        return 'Nezhaarchaeota'
    elif index == 24:
        return 'Verstraetearchaeota'
    elif index == 25:
        return 'Nitrososphaerota'
    elif index == 26:
        return 'Thermoproteota'
    elif index == 27:
        return 'Thermoplasmatota'
    elif index == 28:
        return 'Methanonatronarchaeia'
    elif index == 29:
        return 'Methanofastidiosia'
    elif index == 30:
        return 'Nanohaloarchaea'
    elif index == 31:
        return 'Halobacteria'
    elif index == 32:
        return 'Methanomicrobia'
    else:
        # Just incase there is a weird result
        print (f'Error: Missing Number-{index}')
        print (index==0)

# Just mapping the specific clade with the larger clade
def matchToLargeClade(value:str):
    if value=='Methanococci' or value == 'Methanopyri':
        return 'Methanomada'
    elif value=='Thermococci' or value == 'Theionarchaea':
        return 'Acherontia'   
    elif value=='Hadarchaeia':
        return 'Stygia'   
    elif value=='Odinarchaeota' or value == 'Wukongarchaeota' or value == 'Freyarchaeota' or value == 'Baldrarchaeota' or value == 'Heimdallarchaeota' or value == 'Kariarchaeota' or value == 'Lokiarchaeota' or value == 'Sifarchaeota' or value == 'Thorarchaeota' or value == 'Hodarchaeota' or value == 'Hermodarchaeota' or value=='Helarchaeota' or value=='Brockarchaeota' or value == 'Culexarchaeota' or value == 'Geoarchaeota' or value == 'Korarchaeota' or value == 'Marsarchaeota' or value == 'Verstraetearchaeota' or value == 'Nitrososphaerota' or value == 'Thermoproteota' or value == 'Geothermarchaeota' or value =='Nezhaarchaeota':
        return 'Proteoarchaeaota'   
    elif value=='Thermoplasmatota':
        return 'Diaforarchaea'   
    elif value=='Methanonatronarchaeia' or value == 'Methanofastidiosia' or value == 'Nanohaloarchaea' or value == 'Halobacteria' or value == 'Methanomicrobia':
        return 'Methanotecta' 
    else:
        print (f'Error: Missing Group:{value}')


print ('Running Analysis...')
dataAnalysisParser()
print ('Analysis Complete')
